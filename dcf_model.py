"""
Create DCF Excel model with automatic data fetch from Yahoo Finance.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import yfinance as yf
import pandas as pd
import numpy as np

# Company to analyze - change this to any ticker
TICKER = "ITRI"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DCF Model"

# Styles
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
header_font = Font(bold=True, size=12)
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Helper to set cell
def cell(row, col, value=None, bold=False, fill=None, font_size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=font_size)
    if fill:
        c.fill = fill
    c.alignment = Alignment(horizontal='left' if fill == yellow_fill else 'right')
    c.border = thin_border
    return c

# Helper to safely extract value from pandas series
def get_val(series, default=0):
    if series is None or series.empty:
        return default
    try:
        val = series.iloc[0]
        if pd.isna(val):
            return default
        return val
    except:
        return default

# Fetch data from Yahoo Finance
print(f"Henter data for {TICKER}...")
stock = yf.Ticker(TICKER)
info = stock.info

# Get info
company_name = info.get('longName', info.get('shortName', TICKER))
current_price = info.get('currentPrice', info.get('previousClose', 0))
market_cap = info.get('marketCap', 0)
beta = info.get('beta', 1.0)
shares = info.get('sharesOutstanding', 0)

# Get financials - try different approaches
revenue = 0
ebit = 0
tax_rate = 0.25
depreciation = 0
capex = 0

# Try quarterly financials for more recent data
try:
    quarterly = stock.quarterly_financials
    if quarterly is not None and not quarterly.empty:
        print("  Bruger kvartalsregnskaber...")

        # Revenue
        for idx in quarterly.index:
            if 'Total Revenue' in str(idx) or 'Revenue' in str(idx):
                revenue = get_val(quarterly.loc[idx]) / 1e6
                break

        # EBIT / Operating Income
        for idx in quarterly.index:
            if 'Operating Income' in str(idx):
                ebit = get_val(quarterly.loc[idx]) / 1e6
                break

        # Tax
        for idx in quarterly.index:
            if 'Income Tax' in str(idx):
                tax = get_val(quarterly.loc[idx]) / 1e6
                if ebit > 0:
                    tax_rate = min(abs(tax / ebit), 0.4)
                break

except Exception as e:
    print(f"  Kvartalsdata fejlede: {e}")

# Try cashflow
try:
    cf = stock.cashflow
    if cf is not None and not cf.empty:
        # Depreciation
        for idx in cf.index:
            if 'Depreciation' in str(idx):
                depreciation = abs(get_val(cf.loc[idx])) / 1e6
                break

        # CapEx
        for idx in cf.index:
            if 'Capital Expend' in str(idx):
                capex = abs(get_val(cf.loc[idx])) / 1e6
                break
except:
    pass

# Try balance sheet
total_debt = 0
cash = 0
try:
    bs = stock.balance_sheet
    if bs is not None and not bs.empty:
        # Debt
        for idx in bs.index:
            if 'Total Debt' in str(idx):
                total_debt = get_val(bs.loc[idx]) / 1e6
                break

        # Cash
        for idx in bs.index:
            if 'Cash' in str(idx) and 'equivalent' in str(idx).lower():
                cash = get_val(bs.loc[idx]) / 1e6
                break
except:
    pass

# If still no revenue, try to get from info
if revenue == 0:
    revenue = info.get('totalRevenue', 0) / 1e6

# If still no revenue, try earnings
if revenue == 0:
    try:
        revenue = info.get('revenue', 0)
        if revenue:
            revenue = revenue / 1e6
    except:
        pass

# Calculate margins
ebit_margin = ebit / revenue if revenue > 0 else 0.10
dep_pct = depreciation / revenue if revenue > 0 else 0.03
capex_pct = capex / revenue if revenue > 0 else 0.05
nwc_pct = 0.01  # Default

print(f"  Company: {company_name}")
print(f"  Price: ${current_price:.2f}")
print(f"  Market Cap: ${market_cap/1e6:.1f}M")
print(f"  Revenue: ${revenue:.1f}M")
print(f"  EBIT: ${ebit:.1f}M")
print(f"  EBIT Margin: {ebit_margin:.1%}")
print(f"  Debt: ${total_debt:.1f}M")
print(f"  Cash: ${cash:.1f}M")
print(f"  Tax Rate: {tax_rate:.1%}")

# Company Inputs
cell(1, 1, f"DCF VÆRDIANSÆTTELSE - {company_name}", bold=True, font_size=14)
cell(2, 1, "Juster gule celler for at opdatere")

cell(4, 1, "Virksomhedsdata", bold=True, font_size=12)
cell(5, 1, "Ticker")
cell(5, 2, TICKER, fill=yellow_fill)
cell(6, 1, "Aktiekurs (USD)")
cell(6, 2, current_price if current_price else 0, fill=yellow_fill)
cell(7, 1, "Antal aktier (M)")
cell(7, 2, shares/1e6 if shares else 0, fill=yellow_fill)
cell(8, 1, "Market Cap (USD M)")
cell(8, 2, market_cap/1e6 if market_cap else 0, bold=True, fill=light_blue_fill)
cell(9, 1, "Total Gæld (USD M)")
cell(9, 2, total_debt if total_debt else 0, fill=yellow_fill)
cell(10, 1, "Kontanter (USD M)")
cell(10, 2, cash if cash else 0, fill=yellow_fill)
cell(11, 1, "Beta")
cell(11, 2, beta if beta else 1.0, fill=yellow_fill)
cell(12, 1, "Skatteprocent")
cell(12, 2, tax_rate if tax_rate else 0.25, fill=yellow_fill)

# WACC Parameters
cell(14, 1, "WACC PARAMETRE", bold=True, font_size=12)
cell(15, 1, "Risikofri rente")
cell(15, 2, 0.04, fill=yellow_fill)
cell(16, 1, "Markedsrisikopræmie")
cell(16, 2, 0.055, fill=yellow_fill)
cell(17, 1, "Gældsomkostning (før skat)")
cell(17, 2, 0.05, fill=yellow_fill)
cell(18, 1, "Terminal vækstrate")
cell(18, 2, 0.02, fill=yellow_fill)

# Calculated WACC
cell(19, 1, "WACC", bold=True)
cell(19, 2, "=((B8/(B8+B9))*(B15+B11*B16))+(B9/(B8+B9))*B17*(1-B12))", bold=True, fill=light_blue_fill)

# Revenue forecast
cell(21, 1, "REVENUE PROGNOSE", bold=True, font_size=12)
cell(22, 1, "År")
for i, year in enumerate([2025, 2026, 2027, 2028, 2029]):
    cell(23, 2+i, year)
cell(22, 7, "Gennemsnit")

cell(23, 1, "Revenue vækst")
cell(24, 1, "EBIT margin")
cell(25, 1, "D&A % af revenue")
cell(26, 1, "CapEx % af revenue")
cell(27, 1, "NWC % af revenue")
cell(28, 1, "Skattesats")

cell(23, 7, "=AVERAGE(C23:G23)", fill=light_blue_fill)
cell(24, 7, "=AVERAGE(C24:G24)", fill=light_blue_fill)
cell(25, 7, "=AVERAGE(C25:G25)", fill=light_blue_fill)
cell(26, 7, "=AVERAGE(C26:G26)", fill=light_blue_fill)
cell(27, 7, "=AVERAGE(C27:G27)", fill=light_blue_fill)
cell(28, 7, "=AVERAGE(C28:G28)", fill=light_blue_fill)

# Input assumptions - with some default fades
base_growth = 0.05
for i in range(5):
    cell(23, 2+i, max(base_growth - i*0.005, 0.01), fill=yellow_fill)  # Fade toward 1%
    cell(24, 2+i, max(ebit_margin, 0.05), fill=yellow_fill)
    cell(25, 2+i, max(dep_pct, 0.02), fill=yellow_fill)
    cell(26, 2+i, max(capex_pct, 0.03), fill=yellow_fill)
    cell(27, 2+i, nwc_pct, fill=yellow_fill)
    cell(28, 2+i, tax_rate, fill=yellow_fill)

# Base revenue
cell(29, 1, "Base revenue (USD M)")
cell(29, 2, max(revenue, 100), fill=yellow_fill)

# Calculations
cell(31, 1, "BEREGNINGER", bold=True, font_size=12)
cell(32, 1, "År")
for i, year in enumerate([2025, 2026, 2027, 2028, 2029]):
    cell(32, 2+i, year)

cell(33, 1, "Revenue")
cell(34, 1, "EBIT")
cell(35, 1, "Skat")
cell(36, 1, "NOPAT")
cell(37, 1, "D&A")
cell(38, 1, "CapEx")
cell(39, 1, "NWC ændring")
cell(40, 1, "Free Cash Flow")

# Formulas
for i in range(5):
    col = get_column_letter(3 + i)
    if i == 0:
        ws[f"{col}33"] = f"=B29*(1+C23)"
    else:
        prev_col = get_column_letter(2 + i)
        ws[f"{col}33"] = f"={prev_col}33*(1+{col}23)"
    ws[f"{col}34"] = f"={col}33*{col}24"
    ws[f"{col}35"] = f"={col}34*{col}28"
    ws[f"{col}36"] = f"={col}34-{col}35"
    ws[f"{col}37"] = f"={col}33*{col}25"
    ws[f"{col}38"] = f"={col}33*{col}26"
    ws[f"{col}39"] = f"={col}33*{col}27"
    ws[f"{col}40"] = f"={col}36+{col}37-{col}38-{col}39"

# Terminal Value
cell(41, 1, "Terminal Value")
cell(41, 2, "=G40*(1+B18)/(B19-B18)", bold=True, fill=light_blue_fill)

# Discount factors
cell(42, 1, "Discount factor")
for i in range(5):
    col = get_column_letter(3 + i)
    ws[f"{col}42"] = f"=(1+B19)^{i+1}"

# Discounted FCF
cell(43, 1, "Discounted FCF")
for i in range(5):
    col = get_column_letter(3 + i)
    ws[f"{col}43"] = f"={col}40/{col}42"

# Discounted Terminal Value
cell(44, 1, "Discounted TV")
cell(44, 2, "=B41/(1+B19)^5", bold=True, fill=light_blue_fill)

# Valuation
cell(46, 1, "VÆRDIANSÆTTELSE", bold=True, font_size=12)
cell(47, 1, "PV af FCF")
cell(47, 2, "=SUM(C43:G43)", bold=True, fill=light_blue_fill)
cell(48, 1, "PV af Terminal Value")
cell(48, 2, "=B44", bold=True, fill=light_blue_fill)
cell(49, 1, "Enterprise Value")
cell(49, 2, "=B47+B48", bold=True, fill=green_fill)
cell(50, 1, "Net Gæld")
cell(50, 2, "=B9-B10", bold=True)
cell(51, 1, "Equity Value")
cell(51, 2, "=B49-B50", bold=True, fill=green_fill)
cell(52, 1, "Implied pris per aktie")
cell(52, 2, "=B51/B7", bold=True, fill=green_fill)
cell(53, 1, "Upside/Downside")
cell(53, 2, "=B52/B6-1", bold=True, fill=green_fill)

# Format percentages
for row in [12, 15, 16, 17, 18, 23, 24, 25, 26, 27, 28]:
    for col in range(2, 8):
        ws.cell(row=row, column=col).number_format = '0.00%'

# Number format for money
for row in [6, 7, 8, 9, 10, 29, 33, 34, 35, 36, 37, 38, 39, 40, 41, 43, 44, 47, 48, 49, 50, 51, 52]:
    for col in range(2, 8):
        ws.cell(row=row, column=col).number_format = '#,##0.0'

# Column widths
ws.column_dimensions['A'].width = 25
for i in range(2, 8):
    ws.column_dimensions[get_column_letter(i)].width = 14

# Save
filename = f"DCF_{TICKER}.xlsx"
wb.save(filename)
print(f"\n{filename} er oprettet med data for {company_name}!")
print("Åbn filen og juster de gule celler for at se hvordan værdien ændrer sig.")